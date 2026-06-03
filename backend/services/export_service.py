"""
export_service.py — Tạo file Excel lịch trực theo tuần.

Format: A4 ngang (landscape), 8 cột (A-H), tiếng Việt đầy đủ.
- Tuần bình thường: 1 hàng/ngày, cột SP | NV | LÃNH ĐẠO
- Tuần quyết toán: 2 hàng/ngày (main + sub), danh sách NV dài trong 1 ô
"""
from io import BytesIO
from datetime import date, datetime, timedelta
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

WEEKDAY_VI = {0: "T2", 1: "T3", 2: "T4", 3: "T5", 4: "T6", 5: "T7", 6: "CN"}

# ── Màu sắc ───────────────────────────────────────────────────
CLR_TITLE_BG   = "1F4E79"   # xanh đậm
CLR_TITLE_FG   = "FFFFFF"   # trắng
CLR_HEADER_BG  = "BDD7EE"   # xanh nhạt
CLR_SETTLE_BG  = "EDE7F6"   # tím nhạt (main shift)
CLR_SUB_BG     = "F5F5F5"   # xám nhạt (sub shift)
CLR_WHITE      = "FFFFFF"

# ── Borders ───────────────────────────────────────────────────
_thin = Side(style="thin")
BORDER_ALL = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_NONE = Border()

# Số cột trong file xuất
_NCOLS = 8   # A-H


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color="000000", italic=False, name="Times New Roman") -> Font:
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _apply_row(ws, row: int, values: list, bold=False, fill_hex=None,
               border=True, font_colors=None, size=11, wrap=True):
    """Ghi 1 hàng (values: list _NCOLS phần tử cho cột A-H). Bỏ qua MergedCell."""
    from openpyxl.cell.cell import MergedCell
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx)
        if isinstance(cell, MergedCell):
            continue
        cell.value = val if val != "" else None
        fc = (font_colors[col_idx - 1] if font_colors and col_idx - 1 < len(font_colors)
              else "000000")
        cell.font = _font(bold=bold, size=size, color=fc)
        cell.alignment = _align(wrap=wrap)
        if fill_hex:
            cell.fill = _fill(fill_hex)
        if border:
            cell.border = BORDER_ALL


_SHIFT_TYPE_SUFFIX = {   # N8: label ngày đặc biệt trong cột THỨ
    "cutoff":           " (C/O)",
    "friday":           " (T6)",
    "settlement_main":  " (QT)",
    "settlement_sub":   " (QT-P)",
}


def build_week_excel(shifts: list, week_start: date, week_end: date,
                     signer_name: str = "Nguyễn Quốc Hùng") -> bytes:
    """
    Tạo file .xlsx cho 1 tuần — khổ A4 ngang.
    shifts: list shift dict (từ schedule_service.get_shifts_for_week / _enrich_shift)
    signer_name: N1 — tên người ký từ shift_config (fallback hard-code)
    Trả về bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Tuan {week_start:%d%m}"

    # ── Khổ A4 ngang ──────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9          # A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75,
        header=0.3, footer=0.3,
    )

    # ── Độ rộng cột (8 cột) ───────────────────────────────────
    ws.column_dimensions["A"].width = 6    # THỨ
    ws.column_dimensions["B"].width = 14   # NGÀY
    ws.column_dimensions["C"].width = 28   # SP
    ws.column_dimensions["D"].width = 36   # TẤT CẢ NV (nhiều tên, xuống dòng)
    ws.column_dimensions["E"].width = 38   # LÃNH ĐẠO
    ws.column_dimensions["F"].width = 4    # trống
    ws.column_dimensions["G"].width = 4    # trống
    ws.column_dimensions["H"].width = 4    # trống

    # Phát hiện tuần quyết toán
    is_settlement_week = any(
        s.get("shift_type") in ("settlement_main", "settlement_sub")
        for s in shifts
    )

    # ── Build shift lookup: date_str → list[shift] ────────────
    shift_by_date: dict[str, list] = {}
    for s in shifts:
        d = s.get("shift_date", "")
        shift_by_date.setdefault(d, []).append(s)

    # ── ROW 1: Tiêu đề ────────────────────────────────────────
    title = (f"LỊCH TRỰC TỪ NGÀY {week_start:%d/%m/%Y} "
             f"ĐẾN NGÀY {week_end:%d/%m/%Y}")
    ws.merge_cells("A1:H1")
    ws["A1"].value = title
    ws["A1"].font = _font(bold=True, size=13, color=CLR_TITLE_FG)
    ws["A1"].fill = _fill(CLR_TITLE_BG)
    ws["A1"].alignment = _align(h="center", v="center")
    ws.row_dimensions[1].height = 24

    # ── ROW 2 ─────────────────────────────────────────────────
    if is_settlement_week:
        header_row = 2
    else:
        ws.row_dimensions[2].height = 6
        header_row = 3

    # ── Header ────────────────────────────────────────────────
    ws.merge_cells(f"C{header_row}:D{header_row}")
    _apply_row(ws, header_row,
               ["THỨ", "NGÀY", "NHÂN VIÊN", "", "LÃNH ĐẠO", "", "", ""],
               bold=True, fill_hex=CLR_HEADER_BG, size=11)
    for col in range(1, _NCOLS + 1):
        ws.cell(row=header_row, column=col).alignment = _align(h="center", v="center")
    ws.row_dimensions[header_row].height = 20

    # ── Data rows ─────────────────────────────────────────────
    current_row = header_row + 1
    current = week_start
    while current <= week_end:
        wd = current.weekday()
        if wd >= 5:   # bỏ qua T7, CN
            current += timedelta(days=1)
            continue

        date_str = current.strftime("%Y-%m-%d")
        day_shifts = shift_by_date.get(date_str, [])

        base_thu = WEEKDAY_VI.get(wd, "")
        # N8: thêm suffix loại ca đặc biệt vào cột THỨ
        shift_type_day = (day_shifts[0].get("shift_type", "normal")
                          if day_shifts else "normal")
        thu_label = base_thu + _SHIFT_TYPE_SUFFIX.get(shift_type_day, "")
        date_label = current.strftime("%d/%m/%Y")

        main_shift = next((s for s in day_shifts
                           if s.get("shift_type") != "settlement_sub"), None)
        sub_shift  = next((s for s in day_shifts
                           if s.get("shift_type") == "settlement_sub"), None)

        if main_shift is None and sub_shift is None:
            _apply_row(ws, current_row,
                       [thu_label, date_label, "", "", "", "", "", ""],
                       fill_hex=CLR_WHITE)
            current_row += 1
            current += timedelta(days=1)
            continue

        if is_settlement_week and main_shift:
            # ── Hàng main (quyết toán) ────────────────────────
            main_row_idx = current_row

            leader_name = ""
            if main_shift.get("leader"):
                leader_name = main_shift["leader"].get("full_name", "")
                if main_shift.get("sp_warning") == "leader_sp":
                    leader_name += " (kiêm SP)"

            sp = main_shift.get("sp") or {}
            sp_name = sp.get("full_name", "")
            nvs = main_shift.get("nvs") or []
            nv_names = [nv.get("full_name", "") for nv in nvs]

            all_nv = ([sp_name] if sp_name else []) + nv_names
            all_nv_str = "\n".join(all_nv)

            ws.merge_cells(f"C{current_row}:D{current_row}")
            _apply_row(ws, current_row,
                       [thu_label, date_label, all_nv_str, "", leader_name, "", "", ""],
                       fill_hex=CLR_SETTLE_BG, wrap=True)
            ws.row_dimensions[current_row].height = max(30, 15 * max(1, len(all_nv)))
            current_row += 1

            # ── Hàng sub (quyết toán) ─────────────────────────
            sub_row_idx = current_row

            if sub_shift:
                sub_nvs = sub_shift.get("nvs") or []
                sub_sp = sub_shift.get("sp") or {}
                sub_sp_name = sub_sp.get("full_name", "")
                sub_nv_names = [nv.get("full_name", "") for nv in sub_nvs]
                all_sub = ([sub_sp_name] if sub_sp_name else []) + sub_nv_names
                mid = (len(all_sub) + 1) // 2
                sub_col_c = "\n".join(all_sub[:mid])
                sub_col_d = "\n".join(all_sub[mid:])
                _apply_row(ws, current_row,
                           ["", "", sub_col_c, sub_col_d, "", "", "", ""],
                           fill_hex=CLR_SUB_BG, wrap=True)
                ws.row_dimensions[current_row].height = max(18, 15 * max(1, mid))
            else:
                _apply_row(ws, current_row,
                           ["", "", "", "", "", "", "", ""],
                           fill_hex=CLR_SUB_BG)
                ws.row_dimensions[current_row].height = 18

            # Merge dọc THỨ và NGÀY qua 2 hàng (main + sub)
            ws.merge_cells(f"A{main_row_idx}:A{sub_row_idx}")
            ws.merge_cells(f"B{main_row_idx}:B{sub_row_idx}")
            ws.cell(row=main_row_idx, column=1).alignment = _align(h="center", v="center")
            ws.cell(row=main_row_idx, column=2).alignment = _align(h="center", v="center")
            current_row += 1

        else:
            # ── Hàng bình thường ──────────────────────────────
            shift = main_shift
            leader_name = ""
            nv_col_c = ""
            nv_col_d = ""
            nv_count = 1

            if shift:
                sp_warning = shift.get("sp_warning")
                if shift.get("leader"):
                    leader_name = shift["leader"].get("full_name", "")
                    if sp_warning == "leader_sp":
                        leader_name += " (kiêm SP)"
                sp = shift.get("sp") or {}
                sp_name = sp.get("full_name", "")
                nvs = shift.get("nvs") or []
                nv_names = [nv.get("full_name", "") for nv in nvs]

                nv_col_c = sp_name                   # Cột C: luôn là SP
                nv_col_d = "\n".join(nv_names)        # Cột D: tất cả NV, xuống dòng
                nv_count = max(1, len(nv_names))

            _apply_row(ws, current_row,
                       [thu_label, date_label, nv_col_c, nv_col_d, leader_name, "", "", ""],
                       wrap=True)
            ws.row_dimensions[current_row].height = max(20, 15 * nv_count)
            ws.cell(row=current_row, column=1).alignment = _align(h="center", v="center")
            ws.cell(row=current_row, column=2).alignment = _align(h="center", v="center")
            current_row += 1

        current += timedelta(days=1)

    # ── Ghi chú ───────────────────────────────────────────────
    current_row += 1
    note_row = current_row

    if is_settlement_week:
        ws.cell(row=note_row, column=1).value = "Ghi chú :"
        ws.cell(row=note_row, column=1).font = _font(bold=True, size=10)
        ws.cell(row=note_row, column=2).value = (
            "- Cán bộ không trực chính làm việc theo giờ của hệ thống là 19h"
        )
        ws.cell(row=note_row, column=2).font = _font(size=10, italic=True)
        ws.cell(row=note_row, column=2).alignment = _align(wrap=True)
        ws.merge_cells(f"B{note_row}:H{note_row}")
        ws.row_dimensions[note_row].height = 15

        note_row += 1
        ws.cell(row=note_row, column=2).value = (
            "- Cán bộ không có tên trong lịch trực làm việc theo giờ làm việc của Agribank"
        )
        ws.cell(row=note_row, column=2).font = _font(size=10, italic=True)
        ws.cell(row=note_row, column=2).alignment = _align(wrap=True)
        ws.merge_cells(f"B{note_row}:H{note_row}")
        ws.row_dimensions[note_row].height = 15

        note_row += 2
        ws.cell(row=note_row, column=5).value = "GIÁM ĐỐC"
        ws.cell(row=note_row, column=5).font = _font(bold=True, size=11)
        ws.cell(row=note_row, column=5).alignment = _align(h="center")

        note_row += 3
    else:
        ws.cell(row=note_row, column=1).value = "Ghi chú :"
        ws.cell(row=note_row, column=1).font = _font(bold=True, size=10)
        ws.merge_cells(f"A{note_row}:H{note_row}")
        note_row += 5

    # Chữ ký
    ws.cell(row=note_row, column=5).value = signer_name  # N1: từ shift_config
    ws.cell(row=note_row, column=5).font = _font(size=11)
    ws.cell(row=note_row, column=5).alignment = _align(h="center")

    # ── Print area ────────────────────────────────────────────
    ws.print_area = f"A1:H{note_row}"

    # ── Xuất bytes ────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
